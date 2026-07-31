"""Excel-Import des Druckerbestands — siehe docs/SPEC.md Abschnitt 5.

Ablauf: Datei lesen → filtern → Diff gegen die DB berechnen → anzeigen →
erst nach Bestätigung in einer Transaktion schreiben.

Der Import verändert ausschließlich `printer` und `printer_model`.
Bestände und Bewegungen bleiben unberührt.
"""

from __future__ import annotations

import hashlib
import json
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ImportRun, Printer, PrinterModel

# ─────────────────────────── Spalten ─────────────────────────────────
# Die Excel hat einen ZWEIZEILIGEN Kopf: Zeile 1 = Gruppe, Zeile 2 = Feldname.
# Bei 'ID article' und 'Nom' steht der Name nur in Zeile 1, in Zeile 2 ist None.
HEADER_GROUP_ROW = 0
HEADER_NAME_ROW = 1
FIRST_DATA_ROW = 2

REQUIRED_COLUMNS = [
    "ID article",
    "Nom",
    "Marque",
    "Modèle",
    "Groupe de catégories",
    "ID catégorie d'articles",
    "Remplacé",
    "Volé",
    "Hors service",
    "Type de salle",
]

OPTIONAL_COLUMNS = [
    "S/N",
    "Numéro CGIE",
    "Catégorie d'articles",
    "Code entité",
    "Nom annexe",
    "Salle",
    "Adresse IP",
    "Adresse MAC",
    "Mise en service",
    "Fournisseur",
]

PRINTER_GROUP = "Printer"


class ImportError_(Exception):
    """Fachlicher Importfehler mit Klartextmeldung für die Oberfläche."""


# ─────────────────────────── Hilfsfunktionen ─────────────────────────

_TRUE_WORDS = {"true", "vrai", "wahr", "oui", "ja", "yes", "1", "x"}
_FALSE_WORDS = {"false", "faux", "falsch", "non", "nein", "no", "0", ""}


def as_bool(value: Any) -> bool:
    """Robust gegen bool, 'VRAI'/'FAUX', 1/0 und leere Zellen."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().casefold()
    if text in _TRUE_WORDS:
        return True
    if text in _FALSE_WORDS:
        return False
    raise ImportError_(f"Valeur booléenne non reconnue : « {value} »")


def as_text(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", ascii_text.casefold()).strip("-")


def model_slug(marque: str, modele: str) -> str:
    """Identität eines Modells. Fängt 'HL-L5210dn' vs 'HL-L5210DN' ab."""
    return slugify(f"{marque} {modele}")


def sha256_of(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


# ─────────────────────────── Datei lesen ─────────────────────────────


def read_rows(path: Path) -> list[dict[str, Any]]:
    """Liest die Excel und liefert Zeilen als Dict, Schlüssel = Feldname.

    Wirft ImportError_ mit Klartext, wenn eine Pflichtspalte fehlt — lieber ein
    verständlicher Abbruch als ein stiller Fehlimport.
    """
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - openpyxl wirft diverse Typen
        raise ImportError_(f"Fichier illisible : {exc}") from exc

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) <= FIRST_DATA_ROW:
        raise ImportError_("Le fichier ne contient aucune ligne de données.")

    group_row = rows[HEADER_GROUP_ROW]
    name_row = rows[HEADER_NAME_ROW]

    # Feldname aus Zeile 2, ersatzweise aus Zeile 1 (gilt für 'ID article', 'Nom')
    headers: list[str | None] = []
    for i in range(len(name_row)):
        name = as_text(name_row[i]) or as_text(group_row[i] if i < len(group_row) else None)
        headers.append(name)

    index: dict[str, int] = {}
    for i, name in enumerate(headers):
        if name and name not in index:
            index[name] = i

    missing = [c for c in REQUIRED_COLUMNS if c not in index]
    if missing:
        raise ImportError_(
            "Colonnes obligatoires introuvables : "
            + ", ".join(f"« {c} »" for c in missing)
            + ". Le format de l'export a-t-il changé ?"
        )

    wanted = REQUIRED_COLUMNS + [c for c in OPTIONAL_COLUMNS if c in index]
    out: list[dict[str, Any]] = []
    for raw in rows[FIRST_DATA_ROW:]:
        if all(cell is None for cell in raw):
            continue
        out.append({col: (raw[index[col]] if index[col] < len(raw) else None) for col in wanted})
    return out


# ─────────────────────────── Filter und Diff ─────────────────────────


@dataclass
class PrinterRow:
    """Eine gefilterte, normalisierte Zeile."""

    source_item_id: int
    nom: str
    serial: str | None
    cgie: str | None
    marque: str
    modele: str
    slug: str
    categorie: str | None
    categorie_id: int | None
    code_entite: str | None
    annexe: str | None
    salle: str | None
    salle_type: str | None
    statut: str
    ip: str | None
    mac: str | None
    date_mise_service: date | None
    fournisseur: str | None


@dataclass
class Change:
    printer_nom: str
    champ: str
    avant: str | None
    apres: str | None


@dataclass
class ImportPlan:
    rows_total: int = 0
    rows_kept: int = 0
    rows_state_filtered: int = 0
    rows_category_filtered: int = 0
    category_filtered_detail: dict[str, int] = field(default_factory=dict)
    kept: list[PrinterRow] = field(default_factory=list)
    created: list[PrinterRow] = field(default_factory=list)
    updated: list[tuple[PrinterRow, list[Change]]] = field(default_factory=list)
    absent: list[str] = field(default_factory=list)
    new_models: list[str] = field(default_factory=list)
    new_brands: list[str] = field(default_factory=list)
    models_without_consumables: list[str] = field(default_factory=list)

    @property
    def has_changes(self) -> bool:
        return bool(self.created or self.updated or self.absent)


def normalize_row(raw: dict[str, Any]) -> PrinterRow:
    source_id = as_int(raw.get("ID article"))
    if source_id is None:
        raise ImportError_("Ligne sans « ID article » — impossible à identifier.")

    marque = as_text(raw.get("Marque")) or "?"
    modele = as_text(raw.get("Modèle")) or "?"
    salle_type = as_text(raw.get("Type de salle"))

    return PrinterRow(
        source_item_id=source_id,
        nom=as_text(raw.get("Nom")) or str(source_id),
        serial=as_text(raw.get("S/N")),
        cgie=as_text(raw.get("Numéro CGIE")),
        marque=marque,
        modele=modele,
        slug=model_slug(marque, modele),
        categorie=as_text(raw.get("Catégorie d'articles")),
        categorie_id=as_int(raw.get("ID catégorie d'articles")),
        code_entite=as_text(raw.get("Code entité")),
        annexe=as_text(raw.get("Nom annexe")),
        salle=as_text(raw.get("Salle")),
        salle_type=salle_type,
        # 'Warehouse' = noch nicht installiert. Bewusst über den Typ, nicht über
        # den Text 'Entrepôt' (Akzentvergleich ist fragil).
        statut="entrepot" if salle_type == "Warehouse" else "installe",
        ip=as_text(raw.get("Adresse IP")),
        mac=as_text(raw.get("Adresse MAC")),
        date_mise_service=as_date(raw.get("Mise en service")),
        fournisseur=as_text(raw.get("Fournisseur")),
    )


def filter_rows(raw_rows: list[dict[str, Any]], excluded_category_ids: list[int]) -> ImportPlan:
    plan = ImportPlan(rows_total=len(raw_rows))
    excluded = set(excluded_category_ids or [])

    for raw in raw_rows:
        if as_text(raw.get("Groupe de catégories")) != PRINTER_GROUP:
            continue

        if as_bool(raw.get("Remplacé")) or as_bool(raw.get("Volé")) or as_bool(raw.get("Hors service")):
            plan.rows_state_filtered += 1
            continue

        cat_id = as_int(raw.get("ID catégorie d'articles"))
        if cat_id is not None and cat_id in excluded:
            plan.rows_category_filtered += 1
            label = as_text(raw.get("Catégorie d'articles")) or f"ID {cat_id}"
            plan.category_filtered_detail[label] = plan.category_filtered_detail.get(label, 0) + 1
            continue

        plan.kept.append(normalize_row(raw))

    plan.rows_kept = len(plan.kept)
    return plan


_COMPARED_FIELDS = [
    ("nom", "Nom"),
    ("serial", "S/N"),
    ("cgie", "N° CGIE"),
    ("salle", "Salle"),
    ("salle_type", "Type de salle"),
    ("statut", "Statut"),
    ("annexe", "Annexe"),
    ("ip", "Adresse IP"),
    ("mac", "Adresse MAC"),
    ("fournisseur", "Fournisseur"),
]


def build_plan(session: Session, raw_rows: list[dict[str, Any]], excluded_category_ids: list[int]) -> ImportPlan:
    """Filtert und vergleicht gegen den Datenbestand, ohne zu schreiben."""
    plan = filter_rows(raw_rows, excluded_category_ids)

    existing = {p.source_item_id: p for p in session.scalars(select(Printer)).all()}
    known_models = {m.slug: m for m in session.scalars(select(PrinterModel)).all()}
    known_brands = {m.marque.casefold() for m in known_models.values()}

    seen_slugs: set[str] = set()
    for row in plan.kept:
        if row.slug not in known_models and row.slug not in seen_slugs:
            seen_slugs.add(row.slug)
            plan.new_models.append(f"{row.marque} {row.modele}")
            if row.marque.casefold() not in known_brands:
                if row.marque not in plan.new_brands:
                    plan.new_brands.append(row.marque)

        current = existing.get(row.source_item_id)
        if current is None:
            plan.created.append(row)
            continue

        changes: list[Change] = []
        for attr, label in _COMPARED_FIELDS:
            before = getattr(current, attr)
            after = getattr(row, attr)
            if (before or None) != (after or None):
                changes.append(Change(row.nom, label, before, after))
        model = known_models.get(row.slug)
        if model is None or current.model_id != model.id:
            changes.append(Change(row.nom, "Modèle", current.model.libelle if current.model else None, f"{row.marque} {row.modele}"))
        if current.etat != "actif":
            changes.append(Change(row.nom, "État", current.etat, "actif"))
        if changes:
            plan.updated.append((row, changes))

    kept_ids = {r.source_item_id for r in plan.kept}
    plan.absent = [
        p.nom for p in existing.values() if p.source_item_id not in kept_ids and p.etat == "actif"
    ]

    return plan


# ─────────────────────────── Schreiben ───────────────────────────────


def apply_plan(
    session: Session,
    plan: ImportPlan,
    *,
    filename: str,
    sha256: str,
    user_id: int | None = None,
) -> ImportRun:
    """Schreibt den Plan in einer Transaktion. Löscht nie einen Drucker."""
    run = ImportRun(
        filename=filename,
        sha256=sha256,
        imported_at=datetime.now(),
        user_id=user_id,
        rows_total=plan.rows_total,
        rows_kept=plan.rows_kept,
        rows_state_filtered=plan.rows_state_filtered,
        rows_category_filtered=plan.rows_category_filtered,
        nb_created=len(plan.created),
        nb_updated=len(plan.updated),
        nb_absent=len(plan.absent),
        nb_new_models=len(plan.new_models),
        raw_json=json.dumps(
            [
                {
                    "source_item_id": r.source_item_id,
                    "nom": r.nom,
                    "marque": r.marque,
                    "modele": r.modele,
                    "salle": r.salle,
                    "salle_type": r.salle_type,
                }
                for r in plan.kept
            ],
            ensure_ascii=False,
        ),
    )
    session.add(run)
    session.flush()  # run.id

    models = {m.slug: m for m in session.scalars(select(PrinterModel)).all()}
    printers = {p.source_item_id: p for p in session.scalars(select(Printer)).all()}

    for row in plan.kept:
        model = models.get(row.slug)
        if model is None:
            model = PrinterModel(
                marque=row.marque,
                modele=row.modele,
                slug=row.slug,
                categorie=row.categorie,
                categorie_id=row.categorie_id,
                mapping_ok=0,
            )
            session.add(model)
            session.flush()
            models[row.slug] = model

        printer = printers.get(row.source_item_id)
        if printer is None:
            printer = Printer(source_item_id=row.source_item_id, first_seen_import=run.id)
            session.add(printer)
            printers[row.source_item_id] = printer

        printer.nom = row.nom
        printer.serial = row.serial
        printer.cgie = row.cgie
        printer.model_id = model.id
        printer.code_entite = row.code_entite
        printer.annexe = row.annexe
        printer.salle = row.salle
        printer.salle_type = row.salle_type
        printer.statut = row.statut
        printer.etat = "actif"
        printer.ip = row.ip
        printer.mac = row.mac
        printer.date_mise_service = row.date_mise_service
        printer.fournisseur = row.fournisseur
        printer.last_seen_import = run.id

    # Nicht mehr enthaltene Geräte werden markiert, nicht gelöscht.
    kept_ids = {r.source_item_id for r in plan.kept}
    for printer in printers.values():
        if printer.source_item_id not in kept_ids and printer.etat == "actif":
            printer.etat = "absent"

    session.commit()
    return run
